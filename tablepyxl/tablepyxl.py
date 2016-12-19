# Do imports like python3 so our package works for 2 and 3
from __future__ import absolute_import

from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.utils import column_index_from_string
from premailer import Premailer
from tablepyxl.style import Table


def string_to_int(s):
    if s.isdigit():
        return int(s)
    return 0


def get_Tables(doc):
    soup = BS(doc, "lxml")
    return [Table(table) for table in soup.find_all('table')]


def write_rows(worksheet, elem, row, column=1):
    """
    Writes every tr child element of elem to a row in the worksheet

    returns the next row after all rows are written
    """
    initial_column = column
    for table_row in elem.rows:
        for table_cell in table_row.cells:
            colspan = string_to_int(table_cell.element.get("colspan", "1"))
            rowspan = string_to_int(table_cell.element.get("rowspan", "1"))
            if rowspan > 1 or colspan > 1:
                worksheet.merge_cells(start_row=row, start_column=column,
                                      end_row=row + rowspan - 1, end_column=column + colspan - 1)
            cell = worksheet.cell(row=row, column=column)
            cell.value = table_cell.value
            table_cell.format(cell)
            min_width = table_cell.get_style('min-width') or len(cell.value) + 2
            max_width = table_cell.get_style('max-width')
            width = worksheet.column_dimensions[get_column_letter(column)].width
            if max_width:
                worksheet.column_dimensions[get_column_letter(column)].width = max_width
            elif not width or width < min_width:
                worksheet.column_dimensions[get_column_letter(column)].width = min_width
            column += colspan
        row += 1
        column = initial_column
    return row


def table_to_sheet(table, wb):
    """
    Takes a table and workbook and writes the table to a new sheet.
    The sheet title will be the same as the table attribute name.
    """
    ws = wb.create_sheet(title=table.element.get('name', None))
    insert_table(table, ws, 1, 1)


def document_to_workbook(doc, wb=None, base_url=None):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document.

    The workbook is returned
    """
    if not wb:
        wb = Workbook()
        wb.remove_sheet(wb.active)

    inline_styles_doc = Premailer(doc, base_url=base_url, remove_classes=False).transform()
    tables = get_Tables(inline_styles_doc)

    for table in tables:
        table_to_sheet(table, wb)

    return wb


def document_to_xl(doc, filename, base_url=None):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document. The workbook is written out to a file called filename
    """
    wb = document_to_workbook(doc, base_url=base_url)
    wb.save(filename)


def insert_table(table, worksheet, column, row):
    if table.head:
        row = write_rows(worksheet, table.head, row, column)
    if table.body:
        row = write_rows(worksheet, table.body, row, column)


def insert_table_at_cell(table, cell):
    """
    Inserts a table at the location of an openpyxl Cell object.
    """
    ws = cell.parent
    column, row = cell.column, cell.row
    insert_table(table, ws, column_index_from_string(column), row)
